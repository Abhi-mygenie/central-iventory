#!/usr/bin/env python3
"""Generate POS 4.0 Owner Smoke Batch Excel tracker."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Smoke Batch S-1 to S-19"

# Colors
RED_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FEFCE8", end_color="FEFCE8", fill_type="solid")
GREEN_FILL = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F2937")
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="DC2626")
BODY_FONT = Font(name="Calibri", size=10)
MONEY_FONT = Font(name="Calibri", size=10, bold=True, color="DC2626")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

# Column widths
col_widths = {"A": 6, "B": 28, "C": 10, "D": 75, "E": 12, "F": 30, "G": 18}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# Title row
ws.merge_cells("A1:G1")
ws["A1"] = "POS 4.0 — Owner Smoke Batch (S-1 → S-19)"
ws["A1"].font = TITLE_FONT
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:G2")
ws["A2"] = "Sprint freeze blocked until every item = PASS or owner-attested. | Creds: Qplazm@10 | Preview: https://33cdfa32-2291-4a4d-8438-d55f824b992e.preview.emergentagent.com"
ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="6B7280")
ws.row_dimensions[2].height = 20

# Headers
headers = ["#", "Item", "Priority", "What to Verify", "Result", "Owner Notes", "Date Tested"]
for i, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER
ws.row_dimensions[4].height = 25

# Data
items = [
    # Section: MONEY
    ("section", "🔴 MONEY ITEMS (verify first)"),
    ("S-1", "CR-025 Discount Payload", "P0 💰", "20% discount on ₹1000 order → Network payload order_discount: 200 (₹ amount, NOT 20). self_discount: 0. Repeat on prepaid Place+Pay and Transfer-to-Room.", RED_FILL),
    ("S-10", "BUG-132 Settlement Formula", "P1 💰", "Settlement panel → 6 KPI cards. Expected = Total Funds − Settled (NOT minus pilferage). Pilferage shows backend value (not ₹0). New 'Total Funds' card present. Login: Welcome Resort, pick date with orders.", RED_FILL),
    ("S-11", "CR-039 Credit Total Wire", "P1 💰", "Credit Management → TOTAL CREDIT tile shows real ₹ value (not '—'). TOTAL PAID shows real value. Portfolio export runs instantly (no per-customer API calls). Login: Welcome Resort.", RED_FILL),
    ("S-12", "BUG-133 Check In Filter", "P1 💰", "Any Insights report (Welcome Resort, May dates) → ZERO 'Check In' items anywhere. Previously 118 items = ~₹1.5L phantom revenue. Check all 8 report surfaces: Dashboard, Item Ledger, Order Ledger, Sales, Cancellations, Food Court, Room Orders, Audit Report.", RED_FILL),
    # Section: FEATURE
    ("section", "🟠 FEATURE ITEMS"),
    ("S-2", "CR-018 Schedule Order", "P1", "Schedule checkbox on delivery/takeaway/walk-in (hidden on dine-in/room/QSR). Incomplete time disables Place Order & Collect Bill. Payload: scheduled:1, schedule_at with time. SCH badge on cards. Filter pill. Column order: YTC→Preparing→Ready→Served.", ORANGE_FILL),
    ("S-3", "CR-019 Settings Wizard", "P1", "/restaurant-settings → full 6-step walkthrough → Save & Launch. Values persist on reload.", ORANGE_FILL),
    ("S-4", "CR-020 Phase 4 + B12–B15", "P1", "B11: channel OFF in settings → order-type dropdown hides it. B12: no 'Default GST %' field. B13/B14: 'Item Level / Restaurant Level' labels + dynamic hint. B15: Short Code = toggle, saves 'Yes'/'No'.", ORANGE_FILL),
    ("S-5", "CR-017 WhatsApp Payment Link", "P1", "Unpaid order → OrderCard WhatsApp button → modal pre-fills name/phone → send → customer receives Razorpay link on WhatsApp.", ORANGE_FILL),
    ("S-6", "BUG-116 Realtime Menu Socket", "P1", "Add/edit item from another session/backend → POS menu updates without reload (food_update_${rid}).", ORANGE_FILL),
    ("S-7", "BUG-122 Post-Delivery (3 fixes)", "P1", "POS YTC OrderCard shows ✗ + ✓ (Cancel works, ✓ advances per def_ord_status). Snooze clock ONLY on web-order TableCards.", ORANGE_FILL),
    ("S-8", "BUG-112/113/114", "P1", "112: QSR Place+Pay auto-print fires fast (≤0.5s). 113: split amount fields freely editable, clamp only on blur. 114: category discount → payload carries discount_type + discount_member_category_id/name.", ORANGE_FILL),
    ("S-9", "CR-026 Report Rounding Sweep", "P1", "Reports show paise (₹14.50). Order Ledger has phone/txn-ref/address/room columns. Credit Panel drill-down correct. Bill-summary line order correct, GST/VAT hidden when ₹0.", ORANGE_FILL),
    # Section: PERF/UX
    ("section", "🟡 PERFORMANCE + UX ITEMS"),
    ("S-13", "CR-037 Remove Popular", "P2", "Boot: no 'Popular Items' loading step, no popular-items API call in Network tab. Order Entry: no 'Popular' tab. Boot ~8s faster.", YELLOW_FILL),
    ("S-14", "CR-038 Boot Retry", "P2", "Force boot failure → 'Attempt 1 of 3' counter. After 3 → button disabled + 'Contact support'. Counter is global.", YELLOW_FILL),
    ("S-17", "BUG-131 Sidebar Sticky", "P2", "Expand all sidebar sections → bottom actions (Ringer/Refresh/User/Logout) stay pinned. Don't scroll away.", YELLOW_FILL),
    ("S-18", "CR-044 Insights Cache", "P1 🔒", "Navigate between Insights reports with same date range → no re-fetch (Network tab). Change dates → new call. SECURITY: Logout → login different restaurant → zero cached data from previous.", YELLOW_FILL),
    ("S-19", "CR-045 Field Stripping", "P2", "All Insights reports load correctly with no NaN/missing data. FE-side strip active — backend response unchanged, JS objects trimmed.", YELLOW_FILL),
    # Section: COSMETIC
    ("section", "🟢 COSMETIC ITEMS"),
    ("S-15", "CR-040 Sidebar Labels", "P3", "'Daily Report', 'Daily Summary', 'Daily Room Report'. X/Y/Z Report entries gone. Page headers match.", GREEN_FILL),
    ("S-16", "CR-042 Item Ledger Rename", "P3", "Sidebar + page header: 'Item Ledger' (not 'Items & Menu'). Export title. Audit tab header.", GREEN_FILL),
]

row = 5
for item in items:
    if item[0] == "section":
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"] = item[1]
        ws[f"A{row}"].font = SECTION_FONT
        ws[f"A{row}"].alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1
        continue

    sid, title, priority, verify, fill = item
    ws.cell(row=row, column=1, value=sid).font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=row, column=2, value=title).font = MONEY_FONT if "💰" in priority else BODY_FONT
    ws.cell(row=row, column=3, value=priority).font = MONEY_FONT if "💰" in priority else BODY_FONT
    ws.cell(row=row, column=4, value=verify).font = BODY_FONT
    ws.cell(row=row, column=5, value="☐").font = Font(name="Calibri", size=12)
    ws.cell(row=row, column=6, value="").font = BODY_FONT
    ws.cell(row=row, column=7, value="").font = BODY_FONT

    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = WRAP

    ws.row_dimensions[row].height = 60
    row += 1

# Summary row
row += 1
ws.merge_cells(f"A{row}:G{row}")
ws[f"A{row}"] = "TOTALS: 19 items (4 money · 8 feature · 5 perf/UX · 2 cosmetic) | On full PASS → agent removes DEBUG-B11 logs, flips statuses to CLOSED, cuts baseline, sprint freeze."
ws[f"A{row}"].font = Font(name="Calibri", size=9, italic=True, bold=True, color="374151")

# Data validation for Result column
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(type="list", formula1='"☐,✅ PASS,❌ FAIL,⏸ DEFER"', allow_blank=True)
dv.prompt = "Pick result"
dv.promptTitle = "Smoke Result"
ws.add_data_validation(dv)
for r in range(6, row):
    cell = ws.cell(row=r, column=5)
    if cell.value == "☐":
        dv.add(cell)

# Freeze panes
ws.freeze_panes = "A5"

# Print setup
ws.print_area = f"A1:G{row}"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1

OUTPUT = "/app/frontend/public/downloads/POS4_0_SMOKE_BATCH_TRACKER.xlsx"
import os
os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
wb.save(OUTPUT)
print(f"✅ Saved to {OUTPUT}")
